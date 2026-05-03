#!/usr/bin/env python3
"""
Mr. Kim's Cafe — April 2026 Top 3 Menu Analysis
CFO Financial Insights for May Marketing (Instagram Reels)
Post-Ramadan Iraq Market Context
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Top 3 Menu Analysis"

# Header
ws['A1'] = "Mr. Kim's Cafe — April 2026 Top 3 Menu Analysis"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:H1')

ws['A2'] = "Post-Ramadan Iraq Market | May 2026 Marketing Insights"
ws['A2'].font = Font(italic=True, size=10, color='666666')
ws.merge_cells('A2:H2')

# Assumptions
ws['A4'] = "KEY ASSUMPTIONS (Post-Ramadan Context)"
ws['A4'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A4'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A4:H4')

assumptions = [
    ("Exchange Rate", "IQD/USD", 1310, "CBE official rate May 2026"),
    ("COGS % — Cold Drinks", "%", 32, "Ice, syrup, milk (higher waste)"),
    ("COGS % — Hot Coffee", "%", 28, "Espresso, milk, labor"),
    ("COGS % — Premium", "%", 26, "Vanilla adds perceived value"),
    ("PG Fee (if enabled)", "%", 2.5, "Zain/Switch card fee (NOT ACTIVE YET)"),
    ("Avg Transaction", "IQD", 8500, "2-3 drinks + 1 snack baseline"),
    ("April Working Days", "days", 25, "Ramadan reduced operations"),
    ("Post-Ramadan Lift (May)", "%", 12, "Conservative demand spike"),
]

row = 5
for label, unit, value, note in assumptions:
    ws[f'A{row}'] = label
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = Font(color='0000FF')  # Blue = Input
    ws[f'D{row}'] = note
    row += 1

# TOP 3 MENU SELECTION
ws['A14'] = "TOP 3 MENU SELECTION & RATIONALE"
ws['A14'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A14'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A14:H14')

row = 15
top3 = [
    ("Rank", "Menu Item", "Category", "April Insight", "May Marketing"),
    ("1", "Iced Latte (I002)", "Cold", "High volume post-fasting", "Affordability + cooling"),
    ("2", "Americano (C001)", "Coffee", "Entry price, year-round staple", "Quality + value narrative"),
    ("3", "Vanilla Latte (C005)", "Premium", "High margin, undermarketed", "Premium indulgence angle"),
]

for rank, item, cat, apr, may in top3:
    ws[f'A{row}'] = rank
    ws[f'B{row}'] = item
    ws[f'C{row}'] = cat
    ws[f'D{row}'] = apr
    ws[f'E{row}'] = may
    ws.row_dimensions[row].height = 25
    for col in ['D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    if rank == "Rank":
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    row += 1

# FINANCIAL TABLE
ws['A23'] = "APRIL 2026 FINANCIAL SNAPSHOT (3 Scenarios)"
ws['A23'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A23'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A23:H23')

# 1. ICED LATTE
row = 24
data_i002 = [
    ("ICED LATTE (I002) — #1 Best Seller", "", "", "", ""),
    ("Monthly Units", "units", 420, 520, 680),
    ("Price per Unit", "IQD", 5000, 5000, 5000),
    ("Revenue", "IQD", 2100000, 2600000, 3400000),
    ("COGS %", "%", 32, 32, 32),
    ("COGS per Unit", "IQD", 1600, 1600, 1600),
    ("Total COGS", "IQD", 672000, 832000, 1088000),
    ("Gross Margin $", "IQD", 1428000, 1768000, 2312000),
    ("Gross Margin %", "%", "68%", "68%", "68%"),
]

for metric, unit, cons, base, opt in data_i002:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons
    ws[f'D{row}'] = base
    ws[f'E{row}'] = opt

    if metric in ["Monthly Units", "Price per Unit", "COGS %"]:
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(color='0000FF')

    row += 1

# 2. AMERICANO
row += 1
data_c001 = [
    ("AMERICANO (C001) — #2 Volume Driver", "", "", "", ""),
    ("Monthly Units", "units", 510, 610, 780),
    ("Price per Unit", "IQD", 3500, 3500, 3500),
    ("Revenue", "IQD", 1785000, 2135000, 2730000),
    ("COGS %", "%", 28, 28, 28),
    ("COGS per Unit", "IQD", 980, 980, 980),
    ("Total COGS", "IQD", 499800, 598200, 764400),
    ("Gross Margin $", "IQD", 1285200, 1536800, 1965600),
    ("Gross Margin %", "%", "72%", "72%", "72%"),
]

for metric, unit, cons, base, opt in data_c001:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons
    ws[f'D{row}'] = base
    ws[f'E{row}'] = opt

    if metric in ["Monthly Units", "Price per Unit", "COGS %"]:
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(color='0000FF')

    row += 1

# 3. VANILLA LATTE
row += 1
data_c005 = [
    ("VANILLA LATTE (C005) — #3 Margin Champion", "", "", "", ""),
    ("Monthly Units", "units", 180, 240, 340),
    ("Price per Unit", "IQD", 5000, 5000, 5000),
    ("Revenue", "IQD", 900000, 1200000, 1700000),
    ("COGS %", "%", 26, 26, 26),
    ("COGS per Unit", "IQD", 1300, 1300, 1300),
    ("Total COGS", "IQD", 234000, 312000, 442000),
    ("Gross Margin $", "IQD", 666000, 888000, 1258000),
    ("Gross Margin %", "%", "74%", "74%", "74%"),
]

for metric, unit, cons, base, opt in data_c005:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons
    ws[f'D{row}'] = base
    ws[f'E{row}'] = opt

    if metric in ["Monthly Units", "Price per Unit", "COGS %"]:
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(color='0000FF')

    row += 1

# Summary Stats
row += 2
ws[f'A{row}'] = "COMBINED TOP 3 METRICS"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws.merge_cells(f'A{row}:E{row}')
row += 1

summary = [
    ("Total Revenue (Top 3)", "IQD", "4,785,000", "5,935,000", "7,830,000"),
    ("Total COGS", "IQD", "1,405,800", "1,742,200", "2,294,400"),
    ("Combined Gross Margin $", "IQD", "3,379,200", "4,192,800", "5,535,600"),
    ("Blended Margin %", "%", "70.6%", "70.6%", "70.6%"),
]

for metric, unit, cons, base, opt in summary:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons
    ws[f'D{row}'] = base
    ws[f'E{row}'] = opt
    row += 1

# MARKETING INSIGHTS
row += 2
ws[f'A{row}'] = "MAY INSTAGRAM REELS — 3 Strategic Content Pieces"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

reels = [
    ("REEL #1: ICED LATTE", "Volume Leader", "520 units/month", "Post-fasting recovery drink", "Push: 'Coolest drink in Baghdad — affordability meets taste'"),
    ("REEL #2: AMERICANO", "Margin % Champion", "72% gross margin (HIGHEST)", "Barista skill narrative", "Push: 'Highest-margin customer base — our most profitable espresso'"),
    ("REEL #3: VANILLA LATTE", "UNDERMARKETED", "240→340 units (+40% upside)", "Premium indulgence angle", "Push: 'Limited edition aspiration — treat yourself after Ramadan'"),
]

for reel, focus, metric, narrative, push in reels:
    ws[f'A{row}'] = reel
    ws[f'B{row}'] = focus
    ws[f'C{row}'] = metric
    ws[f'D{row}'] = narrative
    ws[f'E{row}'] = push

    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 40
    row += 1

# PG INTEGRATION WARNING
row += 2
ws[f'A{row}'] = "⚠️ PG INTEGRATION STATUS (Critical for May Launch)"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

pg_notes = [
    ("Current State (April)", "100% cash sales", "No card integration", "All transactions = immediate cash, zero payment delay", "RISK: Manual POS entry prone to data accuracy drift"),
    ("Post-Launch Scenario", "Est. 40% card adoption", "+2.5% fee per transaction", "If 40% of May sales = cards: ~65K IQD monthly fee", "Margin compresses from 70.6% → 69.8% on card portion"),
    ("May Operational Risk", "If soft launch mid-month", "Training + integration friction", "Staff errors, system downtime possible", "Recommendation: Test on low-volume weekdays first"),
]

for state, curr, impact1, impact2, note in pg_notes:
    ws[f'A{row}'] = state
    ws[f'B{row}'] = curr
    ws[f'C{row}'] = impact1
    ws[f'D{row}'] = impact2
    ws[f'E{row}'] = note

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 30
    row += 1

# KPI TARGETS
row += 2
ws[f'A{row}'] = "MAY KPI TARGETS (12% Post-Ramadan Lift)"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws.merge_cells(f'A{row}:E{row}')
row += 1

kpis = [
    ("Metric", "April Base", "May Target (+12%)", "Success Metric", ""),
    ("Total Units (Top 3)", "1,350 units", "1,512 units", "Track daily avg per location", ""),
    ("Total Revenue", "5.9M IQD", "6.6M IQD", "Revenue growth pacing", ""),
    ("Blended Margin", "70.6%", "70.6% (hold)", "Monitor margin dilution", ""),
    ("Reel #1 (I002) Focus", "520 units", "583 units", "Volume KPI — easiest to push", ""),
    ("Reel #2 (C001) Focus", "610 units", "683 units", "Cross-sell food bundles", ""),
    ("Reel #3 (C005) Focus", "240 units", "269 units → 338 units", "Premium push = +40% upside", "Highest ROI on marketing spend"),
]

for metric, apr, may_target, success, note in kpis:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = apr
    ws[f'C{row}'] = may_target
    ws[f'D{row}'] = success
    ws[f'E{row}'] = note

    if metric == "Metric":
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 20
    row += 1

# Format columns
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 30

wb.save('/Users/alli/Desktop/mr.kim\'s claude code/april_2026_top3_analysis.xlsx')
print("✅ Excel file created successfully: april_2026_top3_analysis.xlsx")
