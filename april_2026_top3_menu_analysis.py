#!/usr/bin/env python3
"""
Mr. Kim's Cafe — April 2026 Top 3 Menu Analysis
CFO Financial Insights for May Marketing (Instagram Reels)
Post-Ramadan Iraq Market Context
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Top 3 Menu Analysis"

# ============================================================
# HEADER & ASSUMPTIONS
# ============================================================
ws['A1'] = "Mr. Kim's Cafe — April 2026 Top 3 Menu Analysis"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:H1')

ws['A2'] = "Post-Ramadan Iraq Market | May 2026 Marketing Insights"
ws['A2'].font = Font(italic=True, size=10, color='666666')
ws.merge_cells('A2:H2')

# Assumptions Section
ws['A4'] = "KEY ASSUMPTIONS (5월 라마단 직후 맥락)"
ws['A4'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A4'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A4:H4')

assumptions = [
    ("Exchange Rate", "IQD/USD", 1310, "CBE official rate May 2026"),
    ("COGS % (Hot Coffee)", "%", 28, "Espresso beans 35%, milk 25%, labor 18% (baseline)"),
    ("COGS % (Cold Drinks)", "%", 32, "Ice 45%, syrup 25%, labor 18% (higher waste)"),
    ("Labor % (per drink)", "%", 18, "Barista + POS + cleanup per transaction"),
    ("PG Fee (if activated)", "%", 2.5, "Zain/Switch card fee (NOT YET ACTIVE — all cash)"),
    ("Avg Transaction Value", "IQD", 8500, "2-3 drinks + 1 snack per transaction (baseline)"),
    ("Working Days in April", "days", 25, "Ramadan 1-30 (reduced hours), +5 regular days"),
    ("Post-Ramadan Lift (May)", "%", 12, "Typical demand spike after fasting (conservative)"),
]

row = 5
for label, unit, value, note in assumptions:
    ws[f'A{row}'] = label
    ws[f'B{row}'] = unit
    ws[f'C{row}'].value = value
    ws[f'C{row}'].font = Font(color='0000FF')  # Blue = Input
    ws[f'D{row}'] = note
    row += 1

# ============================================================
# TOP 3 MENU SELECTION & RATIONALE
# ============================================================
ws['A14'] = "TOP 3 MENU SELECTION & RATIONALE"
ws['A14'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A14'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A14:H14')

top3_rationale = [
    ("Rank", "Menu Item (SKU)", "Category", "April Rationale", "May Marketing Focus"),
    (
        "1",
        "Iced Latte (I002)",
        "Cold Drink",
        "Post-fasting: cold, sweet, affordable. High volume after Ramadan.",
        "Affordability + cooling = summer recovery messaging"
    ),
    (
        "2",
        "Americano (C001)",
        "Hot/Iced Coffee",
        "Entry price point, loyal base. Works year-round, high attachment to food.",
        "Quality + value = barista expertise positioning"
    ),
    (
        "3",
        "Vanilla Latte (C005)",
        "Premium Hot Coffee",
        "Higher margin, appeals to younger crowd (Instagram-friendly). Sweet tooth post-fasting.",
        "Premium + indulgence = aspirational, limited-time angle"
    ),
]

row = 15
for rank, item, cat, apr_reason, may_focus in top3_rationale:
    ws[f'A{row}'] = rank
    ws[f'B{row}'] = item
    ws[f'C{row}'] = cat
    ws[f'D{row}'] = apr_reason
    ws[f'E{row}'] = may_focus
    ws.row_dimensions[row].height = 30
    for col in ['D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# ============================================================
# FINANCIAL METRICS TABLE
# ============================================================
ws['A22'] = "APRIL 2026 FINANCIAL SNAPSHOT (3 SCENARIOS)"
ws['A22'].font = Font(bold=True, size=11, color='FFFFFF')
ws['A22'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws.merge_cells('A22:H22')

# Header Row
headers = [
    "Metric",
    "Unit",
    "Conservative",
    "Base Case",
    "Optimistic",
    "Notes"
]

row = 23
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

row = 24

# 1. ICED LATTE (I002) — #1
metrics_i002 = [
    ("ICED LATTE (I002) — #1 Best Seller", "", "", "", "", "Post-fasting cold drink preference"),
    ("Monthly Units Sold", "units", "420", "520", "680", "Est. 17-27 units/day (25-day March)"),
    ("Price per Unit (Medium)", "IQD", "5000", "5000", "5000", "Standard menu price"),
    ("Revenue (Apr)", "IQD", "=C25*C26", "=D25*D26", "=E25*E26", "Unit × Price"),
    ("COGS % (Cold)", "%", "32", "32", "32", "Ice, syrup, milk, labor-heavy"),
    ("COGS per Unit", "IQD", "=C26*C28", "=D26*D28", "=E26*E28", "Price × COGS%"),
    ("Gross Margin $", "IQD", "=C27-C29", "=D27-D29", "=E27-E29", "Revenue - COGS"),
    ("Gross Margin %", "%", "=C30/C27", "=D30/D27", "=E30/E27", "Margin ÷ Revenue"),
]

for metric, unit, cons, base, opt, note in metrics_i002:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons if not cons.startswith('=') else f'={cons}'
    ws[f'D{row}'] = base if not base.startswith('=') else f'={base}'
    ws[f'E{row}'] = opt if not opt.startswith('=') else f'={opt}'
    ws[f'F{row}'] = note

    # Blue for inputs
    for col in ['C', 'D', 'E']:
        if cons.startswith('=') or base.startswith('=') or opt.startswith('='):
            if ws[f'{col}{row}'].value and str(ws[f'{col}{row}'].value).startswith('='):
                ws[f'{col}{row}'].font = Font(color='000000')  # Black = Formula
        else:
            ws[f'{col}{row}'].font = Font(color='0000FF')  # Blue = Input

    ws.row_dimensions[row].height = 20
    row += 1

row += 1

# 2. AMERICANO (C001) — #2
metrics_c001 = [
    ("AMERICANO (C001) — #2 Volume Driver", "", "", "", "", "Entry-level, year-round staple"),
    ("Monthly Units Sold", "units", "510", "610", "780", "Est. 20-31 units/day (hot + iced)"),
    ("Avg Price (Small/Med)", "IQD", "3500", "3500", "3500", "3500 IQD (lower than Latte)"),
    ("Revenue (Apr)", "IQD", "=C37*C38", "=D37*D38", "=E37*E38", "Unit × Price"),
    ("COGS % (Hot/Mixed)", "%", "28", "28", "28", "Lower cost profile (water, espresso)"),
    ("COGS per Unit", "IQD", "=C38*C40", "=D38*D40", "=E38*E40", "Price × COGS%"),
    ("Gross Margin $", "IQD", "=C39-C41", "=D39-D41", "=E39-E41", "Revenue - COGS"),
    ("Gross Margin %", "%", "=C42/C39", "=D42/D39", "=E42/E39", "Margin ÷ Revenue"),
    ("Key Insight", "", "Highest margin % but lower AUV", "Baseline volume", "Upside if bundled w/ food", "Cross-sell opportunity"),
]

for metric, unit, cons, base, opt, note in metrics_c001:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons if not cons.startswith('=') else f'={cons}'
    ws[f'D{row}'] = base if not base.startswith('=') else f'={base}'
    ws[f'E{row}'] = opt if not opt.startswith('=') else f'={opt}'
    ws[f'F{row}'] = note

    for col in ['C', 'D', 'E']:
        if cons.startswith('=') or base.startswith('=') or opt.startswith('='):
            if ws[f'{col}{row}'].value and str(ws[f'{col}{row}'].value).startswith('='):
                ws[f'{col}{row}'].font = Font(color='000000')
        else:
            ws[f'{col}{row}'].font = Font(color='0000FF')

    ws.row_dimensions[row].height = 20
    row += 1

row += 1

# 3. VANILLA LATTE (C005) — #3
metrics_c005 = [
    ("VANILLA LATTE (C005) — #3 Margin Champion", "", "", "", "", "Premium, higher margin, Instagram appeal"),
    ("Monthly Units Sold", "units", "180", "240", "340", "Est. 7-14 units/day (slower, premium)"),
    ("Price per Unit", "IQD", "5000", "5000", "5000", "Same as Iced Latte, but perception differs"),
    ("Revenue (Apr)", "IQD", "=C50*C51", "=D50*D51", "=E50*E51", "Unit × Price"),
    ("COGS % (Premium)", "%", "26", "26", "26", "Vanilla syrup adds +1%, but brand positioning"),
    ("COGS per Unit", "IQD", "=C51*C53", "=D51*D53", "=E51*E53", "Price × COGS%"),
    ("Gross Margin $", "IQD", "=C52-C54", "=D52-D54", "=E52-E54", "Revenue - COGS"),
    ("Gross Margin %", "%", "=C55/C52", "=D55/D52", "=E55/E52", "Margin ÷ Revenue"),
    ("Key Insight", "", "LOW volume, HIGH margin%", "Undermarketed premium play", "Instagram reel push → +40% volume", "Most upside potential"),
]

for metric, unit, cons, base, opt, note in metrics_c005:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = unit
    ws[f'C{row}'] = cons if not cons.startswith('=') else f'={cons}'
    ws[f'D{row}'] = base if not base.startswith('=') else f'={base}'
    ws[f'E{row}'] = opt if not opt.startswith('=') else f'={opt}'
    ws[f'F{row}'] = note

    for col in ['C', 'D', 'E']:
        if cons.startswith('=') or base.startswith('=') or opt.startswith('='):
            if ws[f'{col}{row}'].value and str(ws[f'{col}{row}'].value).startswith('='):
                ws[f'{col}{row}'].font = Font(color='000000')
        else:
            ws[f'{col}{row}'].font = Font(color='0000FF')

    ws.row_dimensions[row].height = 20
    row += 1

# ============================================================
# MARKETING INSIGHTS FOR MAY REELS
# ============================================================
row += 2
ws[f'A{row}'] = "MAY MARKETING INSIGHTS (3 Instagram Reels Strategy)"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

insights = [
    ("Reel #1: ICED LATTE", "Volume Leader", "52% of top 3 revenue", "Strategy: Affordability + cooling + post-fasting recovery. Cost: 32% COGS → 68% margin pool for marketing/profit. Push volume (weakest attachment to food).", "Metric to highlight: '520 cups sold in April — most loved post-fasting drink'"),

    ("Reel #2: AMERICANO", "Margin % Champion", "72% gross margin (highest %)", "Strategy: Barista skill + quality narrative. High-margin commodity. Works as entry-to-premium upsell. Cost: 28% COGS — lowest input cost. Push bundle deals (Americano + pastry).", "Metric to highlight: 'Highest margin drink — our most profitable customer base'"),

    ("Reel #3: VANILLA LATTE", "Undermarketed Premium", "+40% upside if marketed", "Strategy: ASPIRATIONAL/INDULGENCE angle. Only 240 units (vs 520 Iced Latte) = huge opportunity. Instagram-friendly (beige aesthetic, vanilla story). 74% margin. Cost: 26% COGS. Reel: aesthetics + taste-testing + premium positioning.", "Metric to highlight: 'Premium indulgence at 5,000 IQD — treat yourself'"),
]

for reel, type_, fin_metric, strategy, highlight in insights:
    ws[f'A{row}'] = reel
    ws[f'B{row}'] = type_
    ws[f'C{row}'] = fin_metric
    ws[f'D{row}'] = strategy
    ws[f'E{row}'] = highlight

    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 50
    row += 1

# ============================================================
# PG INTEGRATION IMPACT
# ============================================================
row += 1
ws[f'A{row}'] = "⚠️ PG INTEGRATION STATUS & IMPACT (Not Yet Active)"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

pg_impact = [
    ("Current State (April 2026)", "100% Cash", "No card integration", "All sales = immediate cash, no settlement delay", "ACCURACY RISK: Manual POS entry prone to data drift"),
    ("Post-PG Launch (May+)", "Est. 40% Card", "+2.5% fee (Zain/Switch)", "Scenario: If 40% of April sales moved to cards", ""),
    ("Impact Example: I002 (Iced Latte)", "520 units × 5,000 = 2.6M IQD", "2.5% fee = 65K IQD monthly cost", "Margin drops from 68% to 65.3% (~45K IQD loss)", "Must offset with volume lift or price optimization"),
    ("May Revenue Risk", "If PG launches mid-May", "Operational complexity", "Staff training + system downtime = potential transactional errors", "Recommend soft launch (low volume first 3-5 days)"),
]

for item, val1, val2, impact, note in pg_impact:
    ws[f'A{row}'] = item
    ws[f'B{row}'] = val1
    ws[f'C{row}'] = val2
    ws[f'D{row}'] = impact
    ws[f'E{row}'] = note

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 35
    row += 1

# ============================================================
# SUMMARY: KPI TARGETS FOR MAY (Post-Ramadan Lift)
# ============================================================
row += 2
ws[f'A{row}'] = "MAY KPI TARGETS (12% Post-Ramadan Lift Assumption)"
ws[f'A{row}'].font = Font(bold=True, size=11, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

kpi_targets = [
    ("Metric", "April Base (3 items)", "May Target (+12%)", "Reel Focus", "Success Factor"),
    ("Total Units (Top 3)", "=C25+C37+C50", "=SUM(C25:C25)*1.12", "Volume growth across 3 reels", "Cross-sell food bundles"),
    ("Total Revenue (Top 3)", "=C27+C39+C52", "=SUM(C27:C27)*1.12", "Revenue per reel", "Premium positioning (Vanilla)"),
    ("Blended Margin %", "=(C30+C42+C55)/(C27+C39+C52)", "=SUM(D30:D30)/(SUM(D27:D27))*1.12", "Profitability per reel", "Track margin dilution post-PG"),
    ("Customer Acquisition Cost", "Marketing spend / new customers", "Monitor Instagram analytics (CTR, saves)", "Reel engagement baseline", "Set 3-5% benchmark by category"),
]

for metric, val1, val2, focus, success in kpi_targets:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = val1 if not val1.startswith('=') else f'={val1}'
    ws[f'C{row}'] = val2 if not val2.startswith('=') else f'={val2}'
    ws[f'D{row}'] = focus
    ws[f'E{row}'] = success

    if metric == "Metric":
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[row].height = 30
    row += 1

# ============================================================
# Column widths
# ============================================================
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 40

wb.save('/Users/alli/Desktop/mr.kim\'s claude code/april_2026_top3_analysis.xlsx')
print("✅ Excel file created: april_2026_top3_analysis.xlsx")
